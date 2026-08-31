from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field
from openpyxl import load_workbook
from io import BytesIO

from app.database import get_db
from app.models.stock import MovimientoStock
from app.models.producto import Producto
from app.schemas.stock import MovimientoStockCreate, MovimientoStockRespuesta


router = APIRouter(
    prefix="/stock",
    tags=["Stock"]
)


class EntradaMasivaItem(BaseModel):
    codigo: str
    cantidad: int = Field(gt=0)


class EntradaMasiva(BaseModel):
    items: list[EntradaMasivaItem]
    motivo: str = "Ingreso de mercadería"


@router.get("/", response_model=list[MovimientoStockRespuesta])
def listar_movimientos(db: Session = Depends(get_db)):
    return db.query(MovimientoStock).all()


@router.post("/entrada", response_model=MovimientoStockRespuesta)
def entrada_stock(
    movimiento: MovimientoStockCreate,
    db: Session = Depends(get_db)
):
    producto = db.query(Producto).filter(
        Producto.id == movimiento.producto_id
    ).first()

    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    producto.stock += movimiento.cantidad

    nuevo_movimiento = MovimientoStock(
        producto_id=movimiento.producto_id,
        tipo="entrada",
        cantidad=movimiento.cantidad,
        motivo=movimiento.motivo
    )

    db.add(nuevo_movimiento)
    db.commit()
    db.refresh(nuevo_movimiento)
    return nuevo_movimiento


@router.post("/entrada-masiva")
def entrada_stock_masiva(
    datos: EntradaMasiva,
    db: Session = Depends(get_db)
):
    if not datos.items:
        raise HTTPException(status_code=400, detail="No hay productos para ingresar")

    errores = []
    productos_resueltos = []

    for item in datos.items:
        codigo = item.codigo.strip()
        producto = db.query(Producto).filter(
            (Producto.codigo == codigo) | (Producto.codigo_barras == codigo)
        ).first()

        if not producto:
            errores.append(f"Código no encontrado: {codigo}")
            continue

        productos_resueltos.append((producto, item.cantidad))

    if errores:
        raise HTTPException(status_code=400, detail=" | ".join(errores))

    for producto, cantidad in productos_resueltos:
        producto.stock += cantidad
        db.add(MovimientoStock(
            producto_id=producto.id,
            tipo="entrada",
            cantidad=cantidad,
            motivo=datos.motivo.strip() or "Ingreso de mercadería"
        ))

    db.commit()

    return {
        "ok": True,
        "productos_afectados": len(productos_resueltos),
        "total_unidades": sum(cantidad for _, cantidad in productos_resueltos),
        "mensaje": "Entrada masiva registrada correctamente"
    }


@router.post("/importar-excel")
def importar_excel(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    nombre = (archivo.filename or "").lower()
    if not nombre.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="El archivo debe ser Excel .xlsx o .xlsm")

    try:
        contenido = archivo.file.read()
        wb = load_workbook(BytesIO(contenido), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise HTTPException(status_code=400, detail="El Excel está vacío")

    encabezados = [str(x or "").strip().lower() for x in filas[0]]
    indice_codigo = next((i for i, x in enumerate(encabezados) if x in ("codigo", "código", "codigo_barras", "código_barras", "barcode")), None)
    indice_cantidad = next((i for i, x in enumerate(encabezados) if x in ("cantidad", "cant", "stock")), None)

    if indice_codigo is None or indice_cantidad is None:
        raise HTTPException(status_code=400, detail="El Excel debe tener columnas 'codigo' y 'cantidad'")

    items = []
    errores = []
    acumulados = {}

    for numero_fila, fila in enumerate(filas[1:], start=2):
        codigo = str(fila[indice_codigo] or "").strip()
        valor_cantidad = fila[indice_cantidad] if indice_cantidad < len(fila) else None

        if not codigo and valor_cantidad in (None, ""):
            continue
        if not codigo:
            errores.append(f"Fila {numero_fila}: falta código")
            continue

        try:
            cantidad = int(valor_cantidad)
        except (TypeError, ValueError):
            errores.append(f"Fila {numero_fila}: cantidad inválida para {codigo}")
            continue

        if cantidad <= 0:
            errores.append(f"Fila {numero_fila}: la cantidad debe ser mayor a 0 para {codigo}")
            continue

        producto = db.query(Producto).filter(
            (Producto.codigo == codigo) | (Producto.codigo_barras == codigo)
        ).first()

        if not producto:
            errores.append(f"Fila {numero_fila}: código no encontrado {codigo}")
            continue

        if producto.id in acumulados:
            acumulados[producto.id]["cantidad"] += cantidad
        else:
            acumulados[producto.id] = {
                "producto_id": producto.id,
                "codigo": producto.codigo,
                "codigo_barras": producto.codigo_barras,
                "nombre": producto.nombre,
                "stock_actual": int(producto.stock or 0),
                "cantidad": cantidad
            }

    items = list(acumulados.values())

    return {
        "ok": True,
        "items": items,
        "errores": errores,
        "total_unidades": sum(item["cantidad"] for item in items),
        "productos_afectados": len(items)
    }
